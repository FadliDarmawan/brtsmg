"""
Proof-of-concept: parsing multi-sheet BRT Semarang halte data (per-koridor)
lalu merge jadi HalteMaster (deduplicated) dengan antrian review manual
untuk kasus ambigu.

Sumber data: Rekap_Halte_BRT_Ori.xlsx (19 sheet koridor, format tidak seragam)

Alur:
  1. Parse tiap sheet -> KoridorHalte records (baca by header name, bukan posisi kolom)
  2. Normalisasi kategori (trotoar, rambu/halte) + resolusi koordinat final
     (Aktual -> Kor Lain -> Aplikasi, dengan confidence flag)
  3. Validasi (koordinat kosong / di luar bounding box Semarang)
  4. Merge lintas-koridor jadi HalteMaster:
       - jarak < STRICT_M  -> auto-merge
       - STRICT_M <= jarak < REVIEW_M dan nama mirip -> masuk antrian review
       - selain itu -> halte terpisah
  5. Output: halte_master.json, review_queue.json, koridor_halte.json, stats
"""

import json
import math
import re
import difflib
from collections import defaultdict

import openpyxl

SRC = "/mnt/user-data/uploads/Rekap_Halte_BRT_Ori.xlsx"

# Bounding box longgar untuk kota Semarang & sekitarnya (buat validasi sanity)
LAT_MIN, LAT_MAX = -7.25, -6.85
LNG_MIN, LNG_MAX = 110.20, 110.55

STRICT_M = 20      # di bawah ini -> auto merge (dianggap titik fisik sama)
REVIEW_M = 60      # antara STRICT_M dan ini + nama mirip -> antrian review manual
NAME_SIM_THRESHOLD = 0.55

TROTOAR_MAP = {"ya": True, "tidak": False}
RAMBU_MAP = {
    "halte": "halte",
    "rambu": "rambu",
    "tidak ada halte/rambu": "tidak_ada",
}


def norm_txt(v):
    if v is None:
        return None
    return re.sub(r"\s+", " ", str(v)).strip()


def norm_key(v):
    t = norm_txt(v)
    return t.lower() if t else None


def parse_corridor_name(sheet_name):
    """
    '3A Pelabuhan-Kagok (H)' -> ('3A', 'Pelabuhan-Kagok')
    'M Mangkang - Simpang 5 (H)' -> ('M', 'Mangkang - Simpang 5')
    """
    s = sheet_name.strip()
    s = re.sub(r"\(H\)\s*$", "", s).strip()
    m = re.match(r"^([A-Za-z0-9]+)\s+(.+)$", s)
    if not m:
        return None, s
    return m.group(1), m.group(2).strip()


DIRECTION_TAGS = {"utara", "selatan", "barat", "timur"}


def extract_direction_tag(name):
    """Ambil tag arah di dalam kurung siku, misal 'Irigasi [Utara]' -> 'utara'."""
    if not name:
        return None
    m = re.search(r"\[([^\]]*)\]", name)
    if not m:
        return None
    tag = m.group(1).strip().lower()
    return tag if tag in DIRECTION_TAGS else None


def clean_name_for_match(name):
    """Strip suffix arah semacam [Utara]/[Selatan] dan normalisasi spasi/kapital.
    Dipakai HANYA untuk bandingin nama dasar; keputusan merge tetap
    mempertimbangkan direction_tag secara terpisah (lihat merge_halte)."""
    if not name:
        return ""
    t = re.sub(r"\[[^\]]*\]", "", name)
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def haversine_m(lat1, lng1, lat2, lng2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lng2 - lng1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def is_in_bounds(lat, lng):
    return lat is not None and lng is not None and LAT_MIN <= lat <= LAT_MAX and LNG_MIN <= lng <= LNG_MAX


def resolve_coord(row):
    """
    Prioritas: Aktual -> Kor Lain -> Aplikasi
    Return (lat, lng, source, confidence, note)
    """
    candidates = [
        ("aktual", row.get("lat_aktual"), row.get("lng_aktual")),
        ("kor_lain", row.get("lat_korlain"), row.get("lng_korlain")),
        ("aplikasi", row.get("lat_aplikasi"), row.get("lng_aplikasi")),
    ]
    for source, lat, lng in candidates:
        if lat is None or lng is None:
            continue
        if is_in_bounds(lat, lng):
            confidence = "high" if source == "aktual" else ("medium" if source == "kor_lain" else "low")
            return lat, lng, source, confidence, None
        # nilai ada tapi di luar bounding box -> catat lalu coba kandidat berikutnya
    # semua kandidat gagal / di luar bounds
    for source, lat, lng in candidates:
        if lat is not None and lng is not None:
            return lat, lng, source, "invalid", "koordinat di luar bounding box Semarang"
    return None, None, None, "missing", "tidak ada koordinat sama sekali"


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []
    parse_warnings = []

    for sheet_name in wb.sheetnames:
        if "template" in sheet_name.lower():
            continue
        ws = wb[sheet_name]
        kode, nama_koridor = parse_corridor_name(sheet_name)

        header_row = [norm_txt(c.value) for c in ws[1]]
        header_map = {}
        for idx, h in enumerate(header_row):
            if not h:
                continue
            key = h.lower()
            if key.startswith("no") and len(key) <= 3:
                header_map["no"] = idx
            elif "nama halte" in key:
                header_map["nama_halte"] = idx
            elif "lintang" in key and "aplikasi" in key:
                header_map["lat_aplikasi"] = idx
            elif "bujur" in key and "aplikasi" in key:
                header_map["lng_aplikasi"] = idx
            elif "lintang" in key and "aktual" in key:
                header_map["lat_aktual"] = idx
            elif "bujur" in key and "aktual" in key:
                header_map["lng_aktual"] = idx
            elif "lintang" in key and "kor" in key:
                header_map["lat_korlain"] = idx
            elif "bujur" in key and "kor" in key:
                header_map["lng_korlain"] = idx
            elif key.startswith("arah"):
                header_map["arah"] = idx
            elif "fungsi jalan" in key:
                header_map["fungsi_jalan"] = idx
            elif "trotoar" in key:
                header_map["trotoar"] = idx
            elif "rambu" in key:
                header_map["rambu_halte"] = idx
            elif key.startswith("transit"):
                header_map["transit"] = idx

        required = ["no", "nama_halte", "lat_aktual", "lng_aktual"]
        missing_required = [r for r in required if r not in header_map]
        if missing_required:
            parse_warnings.append(
                f"[{sheet_name}] header wajib tidak ditemukan: {missing_required} -> sheet dilewati"
            )
            continue

        def get(row, key):
            idx = header_map.get(key)
            if idx is None:
                return None
            return row[idx].value if idx < len(row) else None

        for row in ws.iter_rows(min_row=2):
            no_val = get(row, "no")
            if no_val is None or norm_txt(no_val) == "":
                continue

            raw = {
                "no": no_val,
                "nama_halte": norm_txt(get(row, "nama_halte")),
                "lat_aplikasi": get(row, "lat_aplikasi"),
                "lng_aplikasi": get(row, "lng_aplikasi"),
                "lat_aktual": get(row, "lat_aktual"),
                "lng_aktual": get(row, "lng_aktual"),
                "lat_korlain": get(row, "lat_korlain"),
                "lng_korlain": get(row, "lng_korlain"),
                "arah": norm_txt(get(row, "arah")),
                "fungsi_jalan": norm_txt(get(row, "fungsi_jalan")),
                "trotoar_raw": norm_txt(get(row, "trotoar")),
                "rambu_raw": norm_txt(get(row, "rambu_halte")),
                "transit_raw": norm_txt(get(row, "transit")),
            }

            lat, lng, coord_source, coord_confidence, coord_note = resolve_coord(raw)

            trotoar = TROTOAR_MAP.get(norm_key(raw["trotoar_raw"]))
            rambu = RAMBU_MAP.get(norm_key(raw["rambu_raw"]))
            transit = TROTOAR_MAP.get(norm_key(raw["transit_raw"]))  # y/n sama pola

            issues = []
            if not raw["nama_halte"]:
                issues.append("nama halte kosong")
            if coord_confidence in ("invalid", "missing"):
                issues.append(coord_note)
            if raw["trotoar_raw"] and trotoar is None:
                issues.append(f"nilai trotoar tak dikenal: '{raw['trotoar_raw']}'")
            if raw["rambu_raw"] and rambu is None:
                issues.append(f"nilai rambu/halte tak dikenal: '{raw['rambu_raw']}'")

            records.append({
                "sheet_name": sheet_name,
                "koridor_kode": kode,
                "koridor_nama": nama_koridor,
                "no_urut": no_val,
                "nama_halte": raw["nama_halte"],
                "nama_halte_match_key": clean_name_for_match(raw["nama_halte"]),
                "direction_tag": extract_direction_tag(raw["nama_halte"]),
                "arah": raw["arah"],
                "fungsi_jalan": raw["fungsi_jalan"],
                "trotoar": trotoar,
                "rambu_halte": rambu,
                "transit": transit,
                "lat": lat,
                "lng": lng,
                "coord_source": coord_source,
                "coord_confidence": coord_confidence,
                "issues": issues,
            })

    return records, parse_warnings


def merge_halte(records):
    """
    Grid-bucket by ~0.003 deg (~300m) untuk batasi jumlah pasangan yang dibandingkan,
    lalu haversine + name similarity buat mutuskan auto-merge / review / terpisah.
    """
    valid = [r for r in records if r["lat"] is not None and r["lng"] is not None]

    def bucket_of(lat, lng):
        return (round(lat / 0.003), round(lng / 0.003))

    buckets = defaultdict(list)
    for i, r in enumerate(valid):
        buckets[bucket_of(r["lat"], r["lng"])].append(i)

    parent = list(range(len(valid)))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    review_queue = []

    def neighbor_bucket_ids(bx, by):
        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                yield (bx + dx, by + dy)

    checked_pairs = set()
    for (bx, by), idxs in buckets.items():
        neighbor_idxs = []
        for nb in neighbor_bucket_ids(bx, by):
            neighbor_idxs.extend(buckets.get(nb, []))
        neighbor_idxs = list(set(neighbor_idxs))

        for a in idxs:
            for b in neighbor_idxs:
                if a >= b:
                    continue
                pair = (a, b)
                if pair in checked_pairs:
                    continue
                checked_pairs.add(pair)

                ra, rb = valid[a], valid[b]
                if ra["sheet_name"] == rb["sheet_name"]:
                    continue  # dalam koridor yang sama, jangan digabung antar-halte

                # Tag arah [Utara]/[Selatan]/dst yang BEDA = platform fisik berbeda
                # di seberang jalan -> jangan pernah digabung, walau jaraknya dekat
                # dan namanya identik.
                if (ra["direction_tag"] and rb["direction_tag"]
                        and ra["direction_tag"] != rb["direction_tag"]):
                    continue

                dist = haversine_m(ra["lat"], ra["lng"], rb["lat"], rb["lng"])
                if dist >= REVIEW_M:
                    continue

                name_sim = difflib.SequenceMatcher(
                    None, ra["nama_halte_match_key"], rb["nama_halte_match_key"]
                ).ratio()

                if dist < STRICT_M:
                    union(a, b)
                elif name_sim >= NAME_SIM_THRESHOLD:
                    review_queue.append({
                        "record_a": {"koridor": ra["sheet_name"], "nama": ra["nama_halte"], "lat": ra["lat"], "lng": ra["lng"]},
                        "record_b": {"koridor": rb["sheet_name"], "nama": rb["nama_halte"], "lat": rb["lat"], "lng": rb["lng"]},
                        "distance_m": round(dist, 1),
                        "name_similarity": round(name_sim, 2),
                        "reason": "jarak dekat & nama mirip, tapi di atas ambang auto-merge",
                    })

    groups = defaultdict(list)
    for i in range(len(valid)):
        groups[find(i)].append(valid[i])

    halte_master = []
    for gid, members in groups.items():
        lat = sum(m["lat"] for m in members) / len(members)
        lng = sum(m["lng"] for m in members) / len(members)
        best_name = max(members, key=lambda m: len(m["nama_halte"] or ""))["nama_halte"]
        halte_master.append({
            "halte_master_id": f"HM{gid:04d}",
            "nama": best_name,
            "lat": round(lat, 6),
            "lng": round(lng, 6),
            "jumlah_koridor": len(set(m["sheet_name"] for m in members)),
            "koridor_terkait": sorted(set(m["sheet_name"] for m in members)),
            "member_count": len(members),
        })

    return halte_master, review_queue


def main():
    records, parse_warnings = parse_workbook(SRC)
    halte_master, review_queue = merge_halte(records)

    invalid_records = [r for r in records if r["coord_confidence"] in ("invalid", "missing")]
    low_conf_records = [r for r in records if r["coord_confidence"] == "low"]
    issue_records = [r for r in records if r["issues"]]

    stats = {
        "total_sheets_parsed": len(set(r["sheet_name"] for r in records)),
        "total_koridor_halte_rows": len(records),
        "total_halte_master_after_merge": len(halte_master),
        "reduction": len(records) - len(halte_master),
        "auto_merged_multi_koridor": sum(1 for h in halte_master if h["jumlah_koridor"] > 1),
        "review_queue_size": len(review_queue),
        "rows_with_invalid_or_missing_coord": len(invalid_records),
        "rows_using_low_confidence_coord_aplikasi": len(low_conf_records),
        "rows_with_any_issue": len(issue_records),
        "parse_warnings": parse_warnings,
    }

    with open("koridor_halte.json", "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    with open("halte_master.json", "w", encoding="utf-8") as f:
        json.dump(halte_master, f, ensure_ascii=False, indent=2)
    with open("review_queue.json", "w", encoding="utf-8") as f:
        json.dump(review_queue, f, ensure_ascii=False, indent=2)
    with open("issue_rows.json", "w", encoding="utf-8") as f:
        json.dump(issue_records, f, ensure_ascii=False, indent=2)
    with open("stats.json", "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

    print(json.dumps(stats, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
